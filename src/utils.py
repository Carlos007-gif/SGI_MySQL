import pandas as pd
import os
from datetime import datetime
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import seaborn as sns
import logging

# Configurar logging para utils
logger = logging.getLogger('DataUtils')

class DataUtils:
    """Clase con utilidades para manejo de datos y generación de reportes"""
    
    @staticmethod
    def exportar_a_excel(datos, columnas, nombre_archivo=None):
        """
        Exporta datos a un archivo Excel con formato profesional
        
        Args:
            datos (list): Lista de tuplas o diccionarios con los datos
            columnas (list): Nombres de las columnas
            nombre_archivo (str): Nombre del archivo (opcional)
        
        Returns:
            str: Ruta del archivo generado
        """
        # Crear directorio si no existe
        os.makedirs('data', exist_ok=True)
        
        # Generar nombre de archivo si no se proporciona
        if not nombre_archivo:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nombre_archivo = f"reporte_inventario_{timestamp}.xlsx"
        
        ruta_completa = os.path.join('data', nombre_archivo)
        
        try:
            # Crear DataFrame
            df = pd.DataFrame(datos, columns=columnas)
            
            # Crear writer de Excel
            with pd.ExcelWriter(ruta_completa, engine='openpyxl') as writer:
                # Exportar datos
                df.to_excel(writer, sheet_name='Reporte', index=False)
                
                # Formatear hoja
                workbook = writer.book
                worksheet = writer.sheets['Reporte']
                
                # Ajustar ancho de columnas
                for idx, column in enumerate(df.columns):
                    max_length = max(df[column].astype(str).map(len).max(), len(column)) + 2
                    worksheet.column_dimensions[chr(65 + idx)].width = min(max_length, 50)
                
                # Agregar información adicional
                summary_sheet = workbook.create_sheet("Resumen")
                summary_sheet.cell(1, 1, "Reporte Generado el:")
                summary_sheet.cell(1, 2, datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
                summary_sheet.cell(2, 1, "Total de Registros:")
                summary_sheet.cell(2, 2, len(df))
                
                # Formatear cabeceras
                from openpyxl.styles import Font, PatternFill, Alignment
                
                # Formato para cabeceras
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                # Aplicar formato a cabeceras
                for col_num, column_title in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.value = column_title
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Ajustar altura de fila de cabecera
                worksheet.row_dimensions[1].height = 25
                
                # Formatear columnas numéricas
                for col_name in ['Precio Unitario', 'Valor Total', 'Cantidad', 'Total Consumido']:
                    if col_name in df.columns:
                        col_idx = df.columns.get_loc(col_name) + 1 # type: ignore[operator]
                        for row in range(2, len(df) + 2):
                            cell = worksheet.cell(row=row, column=col_idx)
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0.00'
                
                # Añadir bordes
                from openpyxl.styles import Border, Side
                thin_border = Border(left=Side(style='thin'), 
                                   right=Side(style='thin'), 
                                   top=Side(style='thin'), 
                                   bottom=Side(style='thin'))
                
                for row in worksheet.iter_rows(min_row=1, max_row=len(df) + 1, max_col=len(df.columns)):
                    for cell in row:
                        cell.border = thin_border
            
            logger.info(f"✅ Reporte exportado exitosamente a {ruta_completa}")
            return ruta_completa
            
        except Exception as e:
            logger.error(f"❌ Error al exportar a Excel: {e}")
            raise
    
    @staticmethod
    def generar_grafico_stock(tipos, cantidades, parent_frame):
        """
        Genera un gráfico de barras con el stock por tipo de producto
        
        Args:
            tipos (list): Tipos de productos (papel, toner, etc.)
            cantidades (list): Cantidades correspondientes
            parent_frame (tk.Frame): Frame donde se mostrará el gráfico
        
        Returns:
            FigureCanvasTkAgg: Canvas con el gráfico listo para mostrar
        """
        try:
            # Configurar estilo
            plt.style.use('seaborn-v0_8')
            sns.set_palette("husl")
            
            # Crear figura
            fig, ax = plt.subplots(figsize=(10, 6), dpi=100)
            
            # Crear gráfico de barras
            colors = ['#3498db', '#2ecc71', '#e74c3c', '#f39c12', '#9b59b6']
            bars = ax.bar(tipos, cantidades, color=colors[:len(tipos)])
            
            # Añadir etiquetas y título
            ax.set_title('Stock por Tipo de Producto', fontsize=16, fontweight='bold', pad=20)
            ax.set_xlabel('Tipo de Producto', fontsize=12)
            ax.set_ylabel('Cantidad en Stock', fontsize=12)
            ax.grid(axis='y', alpha=0.3, linestyle='--')
            
            # Añadir valores sobre las barras
            for bar in bars:
                height = bar.get_height()
                ax.annotate(f'{int(height):,}',
                            xy=(bar.get_x() + bar.get_width() / 2, height),
                            xytext=(0, 3),
                            textcoords="offset points",
                            ha='center', va='bottom', fontweight='bold')
            
            # Ajustar márgenes
            plt.tight_layout()
            
            # Crear canvas para Tkinter
            canvas = FigureCanvasTkAgg(fig, master=parent_frame)
            canvas.draw()
            
            return canvas
            
        except Exception as e:
            logger.error(f"Error al generar gráfico: {e}")
            raise
    
    @staticmethod
    def validar_entrada_numerica(valor):
        """
        Valida que una entrada sea numérica y positiva
        
        Args:
            valor (str): Valor a validar
        
        Returns:
            tuple: (bool, str) - (es_valido, mensaje_error)
        """
        try:
            num = int(valor)
            if num <= 0:
                return False, "El valor debe ser un número positivo"
            return True, ""
        except ValueError:
            return False, "El valor debe ser un número entero válido"
    
    @staticmethod
    def formatear_moneda(valor):
        """
        Formatea un número como moneda
        
        Args:
            valor (float): Valor a formatear
        
        Returns:
            str: Valor formateado como moneda
        """
        try:
            if valor is None or valor == 0:
                return "$0.00"
            return f"${valor:,.2f}"
        except (ValueError, TypeError):
            return "$0.00"
    
    @staticmethod
    def calcular_valor_total_inventario(productos):
        """
        Calcula el valor total del inventario
        
        Args:
            productos (list): Lista de productos con precio y cantidad
        
        Returns:
            float: Valor total del inventario
        """
        valor_total = 0
        for producto in productos:
            precio = producto.get('precio_unitario', 0)
            cantidad = producto.get('cantidad', 0)
            valor_total += precio * cantidad
        return valor_total
    
    @staticmethod
    def generar_reporte_consumo(consumos):
        """
        Genera un reporte de consumo por producto
        
        Args:
            consumos (list): Lista de consumos con nombre y cantidad
        
        Returns:
            dict: Resumen del consumo con total y promedio
        """
        if not consumos:
            return {"total": 0, "promedio": 0, "productos": []}
        
        total = sum(item.get('cantidad', 0) for item in consumos)
        promedio = total / len(consumos) if consumos else 0
        
        return {
            "total": total,
            "promedio": round(promedio, 2),
            "productos": consumos,
            "fecha_generacion": datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        }

# Función auxiliar para validaciones simples
def safe_int_conversion(valor, default=0):
    """Convierte un valor a entero de forma segura"""
    try:
        return int(valor)
    except (ValueError, TypeError):
        return default